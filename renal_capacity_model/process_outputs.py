# Module for processing results into output suitable for users

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from renal_capacity_model.helpers import get_logger
import os
from datetime import datetime

logger = get_logger(__name__)


def create_results_folder(run_start_time: str) -> str:
    """Creates folder results/run_start_time to save results files to

    Args:
        run_start_time (str): Start time of model run

    Returns:
        path_to_results (str): Folder to save model run results
    """
    path_to_results = os.path.join("results", run_start_time)
    if not os.path.exists(path_to_results):
        os.makedirs(path_to_results)
    return path_to_results


def save_result_files(df_to_save: pd.DataFrame, filename: str, path_to_results: str):
    """Saves results files in CSV and parquet formats

    Args:
        df_to_save (pd.DataFrame): Dataframe to save
        filename (str): Name of Dataframe to save
        path_to_results (str): Folder to save results to
    """
    logger.info(f"💾 Saving {df_to_save}")
    df_to_save.to_parquet(os.path.join(path_to_results, filename + ".parquet"))
    df_to_save.to_csv(os.path.join(path_to_results, filename + ".csv"))


def combine_incident_and_prevalent_counts(df: pd.DataFrame) -> pd.DataFrame:
    """The results dataframes separate out counts of prevalence, mortality, and incidence
    by incident and prevalent patients, for easier debugging. We undo this aggregation
    and combine the counts of both incident and prevalent patients, for model results aimed at users.

    Args:
        df (pd.DataFrame): Results dataframe from a model run

    Returns:
        pd.DataFrame: Results dataframe from a model run with the counts for incident and prevalent patients combined
    """
    processed_dfs = []
    for metric in ["prevalence", "mortality", "incidence"]:
        df_filtered = df[df.index.str.contains(metric, case=False, na=False)].copy()
        df_filtered.loc[:, "index"] = df_filtered.index.str.replace(
            r"_?(incident|prevalent)", "", regex=True
        )
        processed_dfs.append(df_filtered.groupby("index").sum())
    return pd.concat(processed_dfs).fillna(0)


def produce_combined_results_for_all_model_runs(
    results_dfs: list[pd.DataFrame],
) -> pd.DataFrame:
    """Combine processed results for all model runs into one dataframe

    Args:
        results_dfs (list[pd.DataFrame]): List of model results dataframes

    Returns:
        pd.DataFrame: DataFrame of all model results combined and processed
    """
    dfs_processed = [combine_incident_and_prevalent_counts(df) for df in results_dfs]
    combined = pd.concat(
        [df.reset_index().assign(model_run=i + 1) for i, df in enumerate(dfs_processed)]
    )
    return (
        combined.set_index(["index", "model_run"])
        .sort_index(axis=0)
        .sort_index(axis=1)
        .fillna(0)
    )


def write_results_to_excel(path_to_excel_file: str, combined_df: pd.DataFrame):
    """Write combined model results from all model runs to Excel file

    Args:
        path_to_excel_file (str): Path to Excel file
        combined_df (pd.DataFrame): Dataframe of all model results combined and processed
    """
    today_date = datetime.now().strftime("%Y%m%d-%H%M")
    wb = load_workbook(path_to_excel_file)
    for outcome in combined_df.index.get_level_values(0).drop_duplicates():
        ws = wb.create_sheet(title=outcome)
        for r in dataframe_to_rows(
            combined_df.loc[outcome].reset_index(), index=False, header=True
        ):
            ws.append(r)
    results_filepath = path_to_excel_file.replace(
        ".xlsx", f"_results_{today_date}.xlsx"
    )
    wb.save(results_filepath)
    logger.info(f"✍️ Excel format model results written to: \n{results_filepath}")
